from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import os
import math

#A Phase 
wb_a_load = load_workbook('4A-GDA-BF-VERT.xlsx')
ws_a_load = wb_a_load.active
wb_a = load_workbook('NOLOAD-GDA-BF-VERT.xlsx')
ws_a = wb_a.active

#B Phase
wb_b_load = load_workbook('4A-GDB-BF-VERT.xlsx')
ws_b_load = wb_b_load.active
wb_b = load_workbook('NOLOAD-GDB-BF-VERT.xlsx')
ws_b = wb_b.active

#C Phase
wb_c_load = load_workbook('4A-GDC-BF-VERT.xlsx')
ws_c_load = wb_c_load.active
wb_c = load_workbook('NOLOAD-GDC-BF-VERT.xlsx')
ws_c = wb_c.active

Frequency = []
dBuV_A = []
dBuV_B = []
dBuV_C = []
dBuV_A_EMA = []
dBuV_B_EMA = []
dBuV_C_EMA = []
L_dBuV_A = []
L_dBuV_B = []
L_dBuV_C = []
L_dBuV_A_EMA = []
L_dBuV_B_EMA = []
L_dBuV_C_EMA = []
delta_A = []
delta_B = [] 
delta_C = []

#A Phase with Load
for rows in ws_a_load.iter_rows(min_row=2, max_row=451, min_col=1, max_col=3, values_only=True):
    #x-axis : Frequency
    Frequency.append(rows[0])
    #y-axis : dBuV
    L_dBuV_A.append(rows[1])

#A Phase without Load
for rows in ws_a.iter_rows(min_row=2, max_row=451, min_col=1, max_col=3, values_only=True):
    #y-axis : dBuV
    dBuV_A.append(rows[1])

#B Phase with Load
for rows in ws_b_load.iter_rows(min_row=2, max_row=451, min_col=1, max_col=3, values_only=True):
    #y-axis : dBuV
    L_dBuV_B.append(rows[1])

#B Phase without Load
for rows in ws_b.iter_rows(min_row=2, max_row=451, min_col=1, max_col=3, values_only=True):
    #y-axis : dBuV
    dBuV_B.append(rows[1])

#C Phase with Load
for rows in ws_c_load.iter_rows(min_row=2, max_row=451, min_col=1, max_col=3, values_only=True):
    #y-axis : dBuV
    L_dBuV_C.append(rows[1])

#C Phase without Load
for rows in ws_c.iter_rows(min_row=2, max_row=451, min_col=1, max_col=3, values_only=True):
    #y-axis : dBuV
    dBuV_C.append(rows[1])

#Start Exponential Smoothing
damping_factor = 0.5
i = 0
def Exponential_Smoothing(alpha,x1, y0):
    filtered_output = (alpha * x1) + ((1 - alpha) * y0)
    return filtered_output

def percentage_change(load, no_load):
    if((load == None) or (no_load == None)):
        change = None
    else:
        change = ((load - no_load)/no_load)*100
    return change

#Calculate the Exponential Moving Average
for i in range (len(Frequency)):
    if ( (i == 0) or (i == 1) ):
        dBuV_A_EMA.append(dBuV_A[0])
        dBuV_B_EMA.append(dBuV_B[0])
        dBuV_C_EMA.append(dBuV_C[0])
        L_dBuV_A_EMA.append(L_dBuV_A[0])
        L_dBuV_B_EMA.append(L_dBuV_B[0])
        L_dBuV_C_EMA.append(L_dBuV_C[0])
    else:
        dBuV_A_EMA.append(0)
        dBuV_B_EMA.append(0)
        dBuV_C_EMA.append(0)
        L_dBuV_A_EMA.append(0)
        L_dBuV_B_EMA.append(0)
        L_dBuV_C_EMA.append(0)
        A_output_no_load = Exponential_Smoothing(damping_factor, dBuV_A[i-1],     dBuV_A_EMA[i-1])
        A_output_load    = Exponential_Smoothing(damping_factor, L_dBuV_A[i-1], L_dBuV_A_EMA[i-1])
        B_output_no_load = Exponential_Smoothing(damping_factor, dBuV_B[i-1],     dBuV_B_EMA[i-1])
        B_output_load    = Exponential_Smoothing(damping_factor, L_dBuV_B[i-1], L_dBuV_B_EMA[i-1])
        C_output_no_load = Exponential_Smoothing(damping_factor, dBuV_C[i-1],     dBuV_C_EMA[i-1])
        C_output_load    = Exponential_Smoothing(damping_factor, L_dBuV_C[i-1], L_dBuV_C_EMA[i-1])
        dBuV_A_EMA[i]   = A_output_no_load
        L_dBuV_A_EMA[i] = A_output_load
        dBuV_B_EMA[i]   = B_output_no_load
        L_dBuV_B_EMA[i] = B_output_load
        dBuV_C_EMA[i]   = C_output_no_load
        L_dBuV_C_EMA[i] = C_output_load

#RMS Average
def RMS_Average(raw_data):
    number_of_samples = 7
    sum_rms = 0
    rms_avg = []
    index = 0
    count = 0
    for t in range(len(Frequency)):
      if(t >= number_of_samples):
        index = t - number_of_samples
        for count in range (number_of_samples):
            sum_rms = sum_rms + (raw_data[index] ** 2)
            index = index + 1 
        rms_avg.append(math.sqrt(sum_rms/number_of_samples))
        count = 0
        sum_rms = 0
      elif(t<number_of_samples):
          rms_avg.append(None)
    return rms_avg       

rms_avg_no_load_A = RMS_Average(dBuV_A)
rms_avg_load_A = RMS_Average(L_dBuV_A)
rms_avg_no_load_B = RMS_Average(dBuV_B)
rms_avg_load_B = RMS_Average(L_dBuV_B)
rms_avg_no_load_C = RMS_Average(dBuV_C)
rms_avg_load_C = RMS_Average(L_dBuV_C)

for f in range (len(Frequency)):
    delta_A.append(percentage_change(rms_avg_load_A[f],rms_avg_no_load_A[f]))
    delta_B.append(percentage_change(rms_avg_load_B[f],rms_avg_no_load_B[f]))
    delta_C.append(percentage_change(rms_avg_load_C[f],rms_avg_no_load_C[f]))

print(delta_A)
#Percentage change in EMI Emission 
fig, axs = plt.subplots(3)
plt.subplots_adjust(hspace = 0.6)
fig.suptitle('Change in EMI')
axs[0].set_title('A Phase Driver')
axs[0].plot(Frequency, delta_A)
axs[0].set_xticks(np.arange(30,200,20))
axs[0].set_yticks(np.arange(-60,250,30))
axs[1].set_title('B Phase Driver')
axs[1].plot(Frequency, delta_B)
axs[1].set_xticks(np.arange(30,200,20))
axs[1].set_yticks(np.arange(-60,250,30))
axs[2].set_title('C Phase Driver')
axs[2].plot(Frequency, delta_C)
axs[2].set_xticks(np.arange(30,200,20))
axs[2].set_yticks(np.arange(-60,250,30))

graph, result = plt.subplots(2)
plt.subplots_adjust(hspace = 0.6)
graph.suptitle('No Load vs Load')
result[0].set_title('Gate Driver A Without Load')
result[0].plot(Frequency, dBuV_A)
result[0].set_xticks(np.arange(30,200,20))
result[1].set_title('Gate Driver A With Load')
result[1].plot(Frequency, L_dBuV_A)
result[1].set_xticks(np.arange(30,200,20))

graphB, resultB = plt.subplots(2)
plt.subplots_adjust(hspace = 0.6)
graphB.suptitle('No Load vs Load')
resultB[0].set_title('Gate Driver B Without Load')
resultB[0].plot(Frequency, dBuV_B)
resultB[0].set_xticks(np.arange(30,200,20))
resultB[1].set_title('Gate Driver B With Load')
resultB[1].plot(Frequency, L_dBuV_B)
resultB[1].set_xticks(np.arange(30,200,20))

graphC, resultC = plt.subplots(2)
plt.subplots_adjust(hspace = 0.6)
graph.suptitle('No Load vs Load')
resultC[0].set_title('Gate Driver C Without Load')
resultC[0].plot(Frequency, dBuV_C)
resultC[0].set_xticks(np.arange(30,200,20))
resultC[1].set_title('Gate Driver C With Load')
resultC[1].plot(Frequency, L_dBuV_C)
resultC[1].set_xticks(np.arange(30,200,20))

compare, GateDriver = plt.subplots(3)
plt.subplots_adjust(hspace = 0.8)
compare.suptitle('Comparison')
GateDriver[0].set_title('Phase A Gate Driver')
GateDriver[0].plot(Frequency, dBuV_A, color = 'b')
GateDriver[0].plot(Frequency, L_dBuV_A, color = 'r')
GateDriver[0].set_xticks(np.arange(30,200,20))
GateDriver[0].set_xlabel("Frequency")
GateDriver[0].set_ylabel("dBuV")
GateDriver[1].set_title('Phase B Gate Driver')
GateDriver[1].plot(Frequency, dBuV_B, color = 'b')
GateDriver[1].plot(Frequency, L_dBuV_B, color = 'r')
GateDriver[1].set_xticks(np.arange(30,200,20))
GateDriver[1].set_xlabel("Frequency")
GateDriver[1].set_ylabel("dBuV")
GateDriver[2].set_title('Phase C Gate Driver')
GateDriver[2].plot(Frequency, dBuV_C, color = 'b')
GateDriver[2].plot(Frequency, L_dBuV_C, color = 'r')
GateDriver[2].set_xticks(np.arange(30,200,20))
GateDriver[2].set_xlabel("Frequency")
GateDriver[2].set_ylabel("dBuV")

avg, rms_avg = plt.subplots(3)
plt.subplots_adjust(hspace = 0.8)
avg.suptitle('Root Mean Square Average')
rms_avg[0].set_title('Phase A Gate Driver')
rms_avg[0].plot(Frequency, rms_avg_no_load_A, color = 'b')
rms_avg[0].plot(Frequency, rms_avg_load_A, color = 'r')
rms_avg[0].set_xticks(np.arange(30,200,20))
rms_avg[0].set_xlabel("Frequency")
rms_avg[0].set_ylabel("dBuV")
rms_avg[1].set_title('Phase B Gate Driver')
rms_avg[1].plot(Frequency, rms_avg_no_load_B, color = 'b')
rms_avg[1].plot(Frequency, rms_avg_load_B, color = 'r')
rms_avg[1].set_xticks(np.arange(30,200,20))
rms_avg[1].set_xlabel("Frequency")
rms_avg[1].set_ylabel("dBuV")
rms_avg[2].set_title('Phase C Gate Driver')
rms_avg[2].plot(Frequency, rms_avg_no_load_C, color = 'b')
rms_avg[2].plot(Frequency, rms_avg_load_C, color = 'r')
rms_avg[2].set_xticks(np.arange(30,200,20))
rms_avg[2].set_xlabel("Frequency")
rms_avg[2].set_ylabel("dBuV")

plt.show()