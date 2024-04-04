from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import os

wb = load_workbook('4A-GDC-BF-VERT.xlsx')
ws = wb.active

ws.insert_rows(1,1)
worksheet = wb.get_sheet_by_name('Sheet1')

worksheet['A1'].value = 'Frequency'
worksheet['C1'].value = 'Exponential Smoothing'
wb.save('4A-GDC-BF-VERT.xlsx')


Frequency = []
dBuV = []
for rows in ws.iter_rows(min_row=2, max_row=451, min_col=1, max_col=3, values_only=True):
    #x-axis : Frequency
    Frequency.append(rows[0])
    #y-axis : dBuV
    dBuV.append(rows[1])

damping_factor = 0.1
dBuV_EMA = []
#Start Calculating Exponential Moving Average
i = 0
r = 2
column = 3
for i in range (len(dBuV)):
    if ((i == 0) or (i == 1)):
        dBuV_EMA.append(dBuV[0])
        input = ws.cell(row = r, column = 3)
        input.value = dBuV_EMA[i]
    else:
        dBuV_EMA.append(0)
        filtered_output = (damping_factor * (dBuV[i-1])) + ((1-damping_factor)*dBuV_EMA[i-1])
        dBuV_EMA[i] = filtered_output
        input = ws.cell(row = r, column = 3)
        input.value = dBuV_EMA[i]
    r = r+1

wb.save('4A-GDC-BF-VERT.xlsx')  